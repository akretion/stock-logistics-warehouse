# Copyright 2020 Akretion
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

from base64 import b64encode
from collections import OrderedDict
from datetime import datetime
from io import BytesIO

import openpyxl

from odoo import models, _
from odoo.exceptions import ValidationError

EXCEL_COLUMNS = [
    "id",
    "name",
    "product_id",
    "location_id",
    "warehouse_id",
    "qty_multiple",
    "product_min_qty",
    "product_max_qty",
]


class StockInventory(models.Model):
    _inherit = "stock.inventory"

    def _generate_orderpoint_vals(self, line, warehouse):
        """ Used for injecting data into """
        ratio_min = self.env["ir.config_parameter"].get_param(
            "orderpoint_initialize_from_inventory_ratio_min", default=0.5
        )
        ratio_max = self.env["ir.config_parameter"].get_param(
            "orderpoint_initialize_from_inventory_ratio_max", default=1
        )
        vals = OrderedDict(
            {
                "id": "",
                "name": line.product_id.name,
                "product_id": line.product_id.id,
                "location_id": warehouse.lot_stock_id.id,
                "warehouse_id": warehouse.id,
                "qty_multiple": 1,
                "product_min_qty": line.product_qty * ratio_min,
                "product_max_qty": line.product_qty * ratio_max,
            }
        )
        return vals

    def _search_orderpoint_vals(self, line, warehouse):
        orderpoint = self.env["stock.warehouse.orderpoint"].search(
            [
                ("product_id", "=", line.product_id.id),
                ("warehouse_id", "=", warehouse.id),
            ]
        )
        result = False
        if orderpoint:
            # generate xmlid
            orderpoint.export_data(["id"])
            result = OrderedDict(
                {
                    "id": orderpoint.get_xml_id()[orderpoint.id],
                    "name": orderpoint.name,
                    "product_id": orderpoint.product_id.id,
                    "location_id": orderpoint.location_id.id,
                    "warehouse_id": orderpoint.warehouse_id.id,
                    "qty_multiple": orderpoint.qty_multiple,
                    "product_min_qty": orderpoint.product_min_qty,
                    "product_max_qty": orderpoint.product_max_qty,
                }
            )
        return result

    def _get_orderpoint_vals(self, line, warehouse):
        return self._search_orderpoint_vals(
            line, warehouse
        ) or self._generate_orderpoint_vals(line, warehouse)

    def _export_orderpoint_xlsx(self, line_vals):
        wb = openpyxl.Workbook()
        sheet = wb.active
        for idx_col, header in enumerate(EXCEL_COLUMNS, start=1):
            sheet.cell(column=idx_col, row=1).value = header
        for idx_row, row in enumerate(line_vals, start=2):
            for idx_col, val_cell in enumerate(row.values(), start=1):
                sheet.cell(column=idx_col, row=idx_row).value = val_cell
        return wb

    def generate_orderpoint_export_xlsx(self):
        """
        1. Search for a warehouse with the same location as current inventory
        2. If an inventory line can be matched to an orderpoint,
           get its values + xmlid
           If it can't be matched, guess its values
        3. Inject values into an importable excel file
        Returns BytesIO
        """
        self.ensure_one()
        warehouse = self.env["stock.warehouse"].search(
            [("lot_stock_id", "=", self.location_id.id)]
        )
        if len(warehouse.ids) != 1:
            raise ValidationError(
                _(
                    "Warehouse configuration error. Stock locations per"
                    " warehouse should be set and mutually unique."
                )
            )
        line_vals = [
            self._get_orderpoint_vals(line, warehouse) for line in self.line_ids
        ]
        workbook = self._export_orderpoint_xlsx(line_vals)
        excel_file = BytesIO()
        workbook.save(excel_file)
        return excel_file

    def button_generate_orderpoint_export_xlsx(self):
        file = self.generate_orderpoint_export_xlsx()
        vals = {
            "filename": (_("Export orderpoint - ")) + str(datetime.today()) + ".xlsx",
            "file": b64encode(file.getvalue()),
        }
        wizard = self.env["generate.orderpoint.xlsx"].create(vals)
        return {
            "name": _("Orderpoint excel"),
            "view_type": "form",
            "view_mode": "form",
            "res_model": "generate.orderpoint.xlsx",
            "type": "ir.actions.act_window",
            "res_id": wizard.id,
        }
