# Copyright 2020 Akretion
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

import base64

import openpyxl

from ..models.stock_inventory import EXCEL_COLUMNS
from io import BytesIO
from odoo.tests import common


class TestStockWarehouseOrderpoint(common.TransactionCase):
    def _check_identical_contents(self, sheet, expected_values):
        for idx_row, row_vals in enumerate(expected_values, start=1):
            for idx_col, cell_val in enumerate(row_vals, start=1):
                self.assertEqual(
                    sheet.cell(column=idx_col, row=idx_row).value,
                    expected_values[idx_row - 1][idx_col - 1],
                )

    def _get_resulting_sheet(self):
        wizard_before = self.env["generate.orderpoint.xlsx"].search([])
        self.inventory.button_generate_orderpoint_export_xlsx()
        wizard = self.env["generate.orderpoint.xlsx"].search(
            [("id", "not in", wizard_before.ids)]
        )
        file = BytesIO(base64.b64decode(wizard.file.decode("utf-8")))
        wb = openpyxl.load_workbook(file)
        return wb.worksheets[0]

    def setUp(self):
        super(TestStockWarehouseOrderpoint, self).setUp()
        self.inventory = self.env.ref("stock.stock_inventory_0")
        self.warehouse = self.env.ref("stock.warehouse0")
        self.location = self.env.ref("stock.stock_location_stock")
        self.inventory_line = self.env.ref("stock.stock_inventory_line_3")
        self.inventory_line_product = self.env.ref("product.product_product_6")
        lines_to_delete = self.inventory.line_ids.filtered(
            lambda r: r.id != self.inventory_line.id
        )
        lines_to_delete.unlink()

    def test_generate_excel_structure(self):
        """
        Check we have correct structure and values for our excel
        """
        sheet = self._get_resulting_sheet()
        expected_vals = [
            EXCEL_COLUMNS,
            [
                None,
                self.inventory_line_product.name,
                self.inventory_line_product.id,
                self.location.id,
                self.warehouse.id,
                1,
                250.00,
                500.00,
            ],
        ]
        self._check_identical_contents(sheet, expected_vals)

    def test_generate_excel_no_extid(self):
        """
        Test the base case: inventory with no matching orderpoints
        """
        sheet = self._get_resulting_sheet()
        expected_vals = [
            ["id"],
            [None],
        ]
        self._check_identical_contents(sheet, expected_vals)

    def test_generate_excel_with_extid(self):
        """
        1. Create orderpoint matching our line
        2. Test that it is exported correctly
        """
        matching_orderpoint = self.env["stock.warehouse.orderpoint"].create(
            {
                "name": "does not matter",
                "product_id": self.inventory_line.product_id.id,
                "location_id": self.location.id,
                "warehouse_id": self.warehouse.id,
                "qty_multiple": 1,
                "product_min_qty": 123,
                "product_max_qty": 456,
            }
        )
        # generate xmlid
        matching_orderpoint.export_data(["id"])
        sheet = self._get_resulting_sheet()
        self.assertIn(
            "__export__.stock_warehouse_orderpoint", sheet.cell(column=1, row=2).value
        )
        self.assertAlmostEqual(sheet.cell(row=2, column=7).value, 123)
        self.assertAlmostEqual(sheet.cell(row=2, column=8).value, 456)
